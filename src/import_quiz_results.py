import argparse
import os
import re
import sqlite3
from datetime import datetime, timezone

from openpyxl import load_workbook

from round_config import ROUND_NAMES

DB_PATH_DEFAULT = os.path.join("data", "quiz_results.db")

EVENT_FILE_PATTERN = re.compile(r"(?P<date>\d{8})_(?P<location>.+)\.(xlsx|xlsm|xltx|xltm)$")


def parse_event_from_filename(filename):
    basename = os.path.basename(filename)
    match = EVENT_FILE_PATTERN.match(basename)
    if not match:
        raise ValueError(
            f"Unable to parse date and location from filename '{basename}'. "
            "Expected format: YYYYMMDD_Location.xlsx"
        )

    event_date = datetime.strptime(match.group("date"), "%Y%m%d").date().isoformat()
    location = match.group("location").replace("_", " ")
    return event_date, location


def normalize_int(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def _is_bonus_cell(cell) -> bool:
    font = getattr(cell, "font", None)
    if font is None:
        return False
    color = getattr(font, "color", None)
    if color is None:
        return False
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return False
    return rgb.upper().endswith("FBBF24")


def _base_round_max_points(round_name: str) -> int:
    if round_name == "Bilderrunde":
        return 10
    if round_name == "Interessantes":
        return 6
    return 5


def create_schema(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS quiz_events (
            id INTEGER PRIMARY KEY,
            event_date TEXT NOT NULL,
            location TEXT NOT NULL,
            source_file TEXT NOT NULL,
            imported_at TEXT NOT NULL
        )
        """
    )
    duplicates = []
    rows = conn.execute(
        "SELECT source_file, id FROM quiz_events ORDER BY source_file, imported_at DESC, id DESC"
    ).fetchall()
    seen = set()
    for source_file, event_id in rows:
        if source_file in seen:
            duplicates.append(event_id)
        else:
            seen.add(source_file)

    for event_id in duplicates:
        conn.execute(
            "DELETE FROM team_scores WHERE team_id IN (SELECT id FROM quiz_teams WHERE event_id = ?)",
            (event_id,),
        )
        conn.execute("DELETE FROM quiz_teams WHERE event_id = ?", (event_id,))
        conn.execute("DELETE FROM quiz_events WHERE id = ?", (event_id,))

    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_quiz_events_source_file ON quiz_events(source_file)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS quiz_teams (
            id INTEGER PRIMARY KEY,
            event_id INTEGER NOT NULL REFERENCES quiz_events(id),
            team_rank INTEGER,
            team_name TEXT NOT NULL,
            puzzle_points INTEGER,
            total INTEGER,
            bonus_round TEXT
        )
        """
    )
    cursor = conn.execute("PRAGMA table_info(quiz_teams)")
    columns = [row[1] for row in cursor.fetchall()]
    if "penalty_points" in columns and "puzzle_points" not in columns:
        conn.execute("ALTER TABLE quiz_teams RENAME COLUMN penalty_points TO puzzle_points")
        columns = [row[1] for row in conn.execute("PRAGMA table_info(quiz_teams)").fetchall()]
    if "bonus_round" not in columns:
        conn.execute("ALTER TABLE quiz_teams ADD COLUMN bonus_round TEXT")

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS team_scores (
            id INTEGER PRIMARY KEY,
            team_id INTEGER NOT NULL REFERENCES quiz_teams(id),
            round_name TEXT NOT NULL,
            points INTEGER
        )
        """
    )
    conn.commit()


def read_quiz_sheet(excel_path):
    workbook = load_workbook(filename=excel_path, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows())
    if not rows:
        raise ValueError(f"Excel workbook '{excel_path}' is empty.")

    header = [str(cell.value).strip() if cell.value is not None else None for cell in rows[0]]
    if "Teamname" not in header:
        raise ValueError(
            "Expected a header row containing 'Teamname'. "
            "The template in data/quizzes should include that column."
        )

    teamname_idx = header.index("Teamname")
    rank_idx = header.index("#") if "#" in header else None
    pp_idx = header.index("PP") if "PP" in header else None
    total_idx = header.index("Gesamt") if "Gesamt" in header else None

    question_indices = []
    for idx, title in enumerate(header):
        if title is None:
            continue
        if isinstance(title, str) and title.isdigit():
            question_indices.append((idx, int(title)))

    raw_teams = []
    for row in rows[1:]:
        if row is None or len(row) <= teamname_idx:
            continue
        team_name = row[teamname_idx].value
        if team_name is None or str(team_name).strip() == "":
            continue

        bonus_round_numbers = []
        team = {
            "team_rank": normalize_int(row[rank_idx].value) if rank_idx is not None else None,
            "team_name": str(team_name).strip(),
            "puzzle_points": normalize_int(row[pp_idx].value) if pp_idx is not None else None,
            "total": normalize_int(row[total_idx].value) if total_idx is not None else None,
            "bonus_round": None,
            "bonus_round_numbers": [],
            "scores": [],
        }

        for idx, question_number in question_indices:
            cell = row[idx]
            value = normalize_int(cell.value) if idx < len(row) else None
            if _is_bonus_cell(cell):
                bonus_round_numbers.append(question_number)
            team["scores"].append((question_number, value))

        team["bonus_round_numbers"] = bonus_round_numbers
        raw_teams.append(team)

    bonus_rounds = {
        bonus_round
        for team in raw_teams
        for bonus_round in team["bonus_round_numbers"]
    }
    if 3 in bonus_rounds and 7 in bonus_rounds:
        raise ValueError(
            "Invalid bonus-round data: both round 3 and round 7 are selected as bonus rounds in this quiz."
        )

    third_round_scores = []
    for team in raw_teams:
        for question_number, points in team["scores"]:
            if question_number == 3 and points is not None:
                third_round_scores.append(points)

    has_third_round_bonus = 3 in bonus_rounds
    has_high_score_in_round_3 = any(points > 5 for points in third_round_scores)
    shifted_round_order = (not has_third_round_bonus) and has_high_score_in_round_3

    shifted_round_names = {
        1: ROUND_NAMES.get(1, "Round 1"),
        2: ROUND_NAMES.get(2, "Round 2"),
        3: ROUND_NAMES.get(7, "Round 7"),
        4: ROUND_NAMES.get(3, "Round 3"),
        5: ROUND_NAMES.get(4, "Round 4"),
        6: ROUND_NAMES.get(5, "Round 5"),
        7: ROUND_NAMES.get(6, "Round 6"),
        8: ROUND_NAMES.get(8, "Round 8"),
        9: ROUND_NAMES.get(9, "Round 9"),
        10: ROUND_NAMES.get(10, "Round 10"),
    }

    round_name_map = shifted_round_names if shifted_round_order else ROUND_NAMES

    teams = []
    validation_errors = []
    for team in raw_teams:
        team_bonus_round_numbers = team["bonus_round_numbers"]
        if len(team_bonus_round_numbers) > 1:
            selected = ", ".join(str(round_no) for round_no in sorted(team_bonus_round_numbers))
            validation_errors.append(
                f"Team '{team['team_name']}' selected more than one bonus round ({selected})."
            )

        bonus_round_number = team_bonus_round_numbers[0] if len(team_bonus_round_numbers) == 1 else None
        mapped_scores = []
        for question_number, points in team["scores"]:
            round_name = round_name_map.get(question_number, f"Round {question_number}")
            mapped_scores.append((round_name, points))

            if points is None:
                continue

            is_bonus_round = bonus_round_number == question_number
            base_max = _base_round_max_points(round_name)
            allowed_max = base_max + (5 if is_bonus_round else 0)

            if is_bonus_round and round_name == "Bilderrunde":
                validation_errors.append(
                    f"Team '{team['team_name']}' selected 'Bilderrunde' as bonus round, which is not allowed."
                )

            if points > allowed_max:
                bonus_note = " (bonus round)" if is_bonus_round else ""
                validation_errors.append(
                    f"Team '{team['team_name']}' has {points} points in '{round_name}'{bonus_note}; maximum allowed is {allowed_max}."
                )

        team["bonus_round"] = (
            round_name_map.get(bonus_round_number, f"Round {bonus_round_number}")
            if bonus_round_number is not None
            else None
        )
        team["scores"] = mapped_scores
        team.pop("bonus_round_numbers", None)
        teams.append(team)

    if validation_errors:
        details = "\n".join(f"- {message}" for message in validation_errors)
        raise ValueError(
            f"Import validation failed for '{os.path.basename(excel_path)}':\n{details}"
        )

    return teams


def import_quiz_file(excel_path, db_path=DB_PATH_DEFAULT):
    event_date, location = parse_event_from_filename(excel_path)
    teams = read_quiz_sheet(excel_path)

    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    create_schema(conn)

    imported_at = datetime.now(timezone.utc).isoformat()
    cur = conn.cursor()
    source_file = os.path.basename(excel_path)
    cur.execute(
        "SELECT id FROM quiz_events WHERE source_file = ?",
        (source_file,),
    )
    row = cur.fetchone()
    if row is not None:
        event_id = row[0]
        cur.execute(
            "DELETE FROM team_scores WHERE team_id IN (SELECT id FROM quiz_teams WHERE event_id = ?)",
            (event_id,),
        )
        cur.execute("DELETE FROM quiz_teams WHERE event_id = ?", (event_id,))
        cur.execute(
            "UPDATE quiz_events SET event_date = ?, location = ?, imported_at = ? WHERE id = ?",
            (event_date, location, imported_at, event_id),
        )
    else:
        cur.execute(
            "INSERT INTO quiz_events (event_date, location, source_file, imported_at) VALUES (?, ?, ?, ?)",
            (event_date, location, source_file, imported_at),
        )
        event_id = cur.lastrowid

    for team in teams:
        cur.execute(
            "INSERT INTO quiz_teams (event_id, team_rank, team_name, puzzle_points, total, bonus_round) VALUES (?, ?, ?, ?, ?, ?)",
            (
                event_id,
                team["team_rank"],
                team["team_name"],
                team["puzzle_points"],
                team["total"],
                team["bonus_round"],
            ),
        )
        team_id = cur.lastrowid

        for question_number, points in team["scores"]:
            cur.execute(
                "INSERT INTO team_scores (team_id, round_name, points) VALUES (?, ?, ?)",
                (team_id, question_number, points),
            )

    conn.commit()
    conn.close()

    return {
        "event_date": event_date,
        "location": location,
        "teams_imported": len(teams),
        "database": db_path,
    }


def import_quiz_folder(folder_path, db_path=DB_PATH_DEFAULT):
    """Import all Excel quiz files from a folder into the database."""
    if not os.path.isdir(folder_path):
        raise ValueError(f"Folder does not exist: '{folder_path}'")

    excel_extensions = (".xlsx", ".xlsm", ".xltx", ".xltm")
    excel_files = [
        os.path.join(folder_path, name)
        for name in sorted(os.listdir(folder_path))
        if name.lower().endswith(excel_extensions)
    ]

    if not excel_files:
        raise ValueError(f"No Excel files found in folder: '{folder_path}'")

    imported = []
    errors = []
    for excel_path in excel_files:
        try:
            imported.append(import_quiz_file(excel_path, db_path))
        except Exception as exc:
            errors.append(
                {
                    "file": os.path.basename(excel_path),
                    "message": str(exc),
                }
            )

    return {
        "folder": folder_path,
        "files_found": len(excel_files),
        "files_imported": len(imported),
        "files_failed": len(errors),
        "database": db_path,
        "imports": imported,
        "errors": errors,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Import PubQuiz results from an Excel file into a SQLite database."
    )
    parser.add_argument(
        "excel_file",
        nargs="?",
        help="Path to a single quiz result Excel file.",
    )
    parser.add_argument(
        "--folder",
        help="Path to a folder containing quiz result Excel files to import.",
    )
    parser.add_argument(
        "--db",
        default=DB_PATH_DEFAULT,
        help="Path to SQLite database file. Defaults to data/quiz_results.db.",
    )
    args = parser.parse_args()

    if args.folder:
        summary = import_quiz_folder(args.folder, args.db)
        for result in summary["imports"]:
            print(f"Imported {result['teams_imported']} teams from {result['event_date']} @ {result['location']}")

        if summary["errors"]:
            print()
            print("Errors during import:")
            for error in summary["errors"]:
                print(f"\n[{error['file']}]\n{error['message']}")

        print()
        print(f"Imported {summary['files_imported']} files from {summary['folder']}")
        if summary["files_failed"]:
            print(f"Failed: {summary['files_failed']} files")
        print(f"Saved into {summary['database']}")

        if summary["files_failed"]:
            raise SystemExit(1)
        return

    if not args.excel_file:
        parser.error("Provide either 'excel_file' or '--folder'.")

    try:
        result = import_quiz_file(args.excel_file, args.db)
    except Exception as exc:
        print(f"Import failed for '{os.path.basename(args.excel_file)}':")
        print(str(exc))
        raise SystemExit(1)

    print(f"Imported {result['teams_imported']} teams from {result['event_date']} @ {result['location']}")
    print(f"Saved into {result['database']}")


if __name__ == "__main__":
    main()
