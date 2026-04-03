"""Excel file loader."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from pubquizstats.models.schemas import ParsedQuizData, RoundData, TeamScoreInput
from pubquizstats.config import EXCEL_SAMPLE_SIZE

logger = logging.getLogger(__name__)


class ExcelLoader:
    """Loader for Excel quiz files."""

    EXPECTED_HEADERS = {"team", "name", "points", "rank", "score", "total"}

    def __init__(self):
        """Initialize Excel loader."""
        self.parsed_data = None

    def load_quiz(self, file_path: str) -> ParsedQuizData:
        """
        Load quiz from Excel file.
        
        Expects format:
        - First sheet contains quiz metadata and results
        - Column A: Team names
        - Column B onwards: Round scores
        - Last column (or specific column): Rank
        - Last column (or specific column): Total points
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            ParsedQuizData with extracted information
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        if not ws:
            raise ValueError("No active worksheet in Excel file")

        # Extract quiz metadata
        quiz_name = self._extract_quiz_name(file_path, ws)
        quiz_date = self._extract_quiz_date(ws)
        location = self._extract_location(ws)

        # Detect structure
        structure = self._detect_structure(ws)
        logger.info(f"Detected structure: {structure}")

        # Extract rounds and team scores
        rounds, team_scores = self._extract_data(ws, structure)

        # Create ParsedQuizData
        parsed = ParsedQuizData(
            name=quiz_name,
            date=quiz_date,
            location=location,
            rounds=rounds,
            team_scores=team_scores,
        )

        return parsed

    def _extract_quiz_name(self, file_path: Path, ws: Worksheet) -> str:
        """Extract quiz name from metadata or filename."""
        # Try to find quiz name in first few rows
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and "quiz" in cell.lower():
                    return cell.strip()

        # Fallback to filename
        return file_path.stem

    def _extract_quiz_date(self, ws: Worksheet) -> datetime:
        """Extract quiz date from metadata."""
        # Look for date in first few rows
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if isinstance(cell, datetime):
                    return cell
                if isinstance(cell, str) and len(cell) > 0:
                    try:
                        return datetime.fromisoformat(cell)
                    except ValueError:
                        pass

        # Fallback to today
        return datetime.now()

    def _extract_location(self, ws: Worksheet) -> Optional[str]:
        """Extract location from metadata."""
        # Look for location keyword
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and "location" in cell.lower():
                    # Get next cell
                    col = ws.cell(row=row[0], column=2)
                    return col.value

        return None

    def _detect_structure(self, ws: Worksheet) -> Dict[str, Any]:
        """
        Detect Excel structure.
        
        Returns dict with:
        - teams_col: Column letter for team names
        - round_cols: List of (col_letter, round_number, round_name)
        - rank_col: Column letter for ranks
        - total_col: Column letter for total points
        - data_start_row: Row where data starts
        """
        structure = {
            "teams_col": "A",
            "round_cols": [],
            "rank_col": None,
            "total_col": None,
            "data_start_row": None,
        }

        # Find header row (look for common keywords)
        header_row = self._find_header_row(ws)
        structure["data_start_row"] = header_row + 1 if header_row else 2

        if header_row:
            # Parse header row
            headers = []
            for col_idx, cell in enumerate(
                ws[header_row], start=1
            ):
                if cell.value:
                    headers.append((col_idx, str(cell.value).lower()))

            # Identify column types
            round_num = 1
            for col_idx, header_text in headers:
                col_letter = openpyxl.utils.get_column_letter(col_idx)

                if any(kw in header_text for kw in ["team", "name", "group"]):
                    structure["teams_col"] = col_letter
                elif any(kw in header_text for kw in ["rank", "position", "place"]):
                    structure["rank_col"] = col_letter
                elif any(kw in header_text for kw in ["total", "sum", "final"]):
                    structure["total_col"] = col_letter
                elif any(kw in header_text for kw in ["round", "r", "q"]):
                    structure["round_cols"].append(
                        (col_letter, round_num, header_text)
                    )
                    round_num += 1
                elif header_text.isdigit():
                    # Assume numeric headers are round numbers
                    structure["round_cols"].append(
                        (col_letter, int(header_text), f"Round {header_text}")
                    )

        # Fallback: assume standard layout
        if not structure["round_cols"]:
            structure["round_cols"] = [
                (openpyxl.utils.get_column_letter(i), i - 1, f"Round {i - 1}")
                for i in range(2, 10)  # B-J columns
            ]

        return structure

    def _find_header_row(self, ws: Worksheet) -> Optional[int]:
        """Find row containing headers."""
        for row_idx in range(1, min(EXCEL_SAMPLE_SIZE + 1, ws.max_row + 1)):
            row_values = [cell.value for cell in ws[row_idx]]
            # Check if this row looks like a header
            text_cells = sum(
                1 for v in row_values
                if isinstance(v, str)
            )
            if text_cells > 2:  # At least 3 text columns
                return row_idx

        return None

    def _extract_data(
        self, ws: Worksheet, structure: Dict[str, Any]
    ) -> Tuple[list[RoundData], list[TeamScoreInput]]:
        """Extract rounds and team scores from worksheet."""
        rounds = []
        team_scores = []

        # Create RoundData objects
        for col_letter, round_num, round_name in structure["round_cols"]:
            rounds.append(
                RoundData(round_number=round_num, round_name=round_name, max_points=0)
            )

        # Extract team data
        for row_idx in range(
            structure["data_start_row"], ws.max_row + 1
        ):
            team_name_cell = ws[f"{structure['teams_col']}{row_idx}"]
            team_name = team_name_cell.value

            # Skip empty rows
            if not team_name or (isinstance(team_name, str) and not team_name.strip()):
                continue

            team_name = str(team_name).strip()

            # Get rank if available
            rank = None
            if structure["rank_col"]:
                rank_cell = ws[f"{structure['rank_col']}{row_idx}"]
                rank = rank_cell.value
                if rank:
                    try:
                        rank = int(rank)
                    except (ValueError, TypeError):
                        rank = None

            # Get total points
            total_points = 0
            if structure["total_col"]:
                total_cell = ws[f"{structure['total_col']}{row_idx}"]
                total_points = total_cell.value or 0
            else:
                # Calculate from round scores
                for col_letter, _, _ in structure["round_cols"]:
                    cell_value = ws[f"{col_letter}{row_idx}"].value
                    if cell_value:
                        try:
                            total_points += int(cell_value)
                        except (ValueError, TypeError):
                            pass

            # Get round scores
            round_scores = {}
            for col_letter, round_num, _ in structure["round_cols"]:
                cell_value = ws[f"{col_letter}{row_idx}"].value
                if cell_value:
                    try:
                        round_scores[round_num] = int(cell_value)
                    except (ValueError, TypeError):
                        pass

            team_score = TeamScoreInput(
                team_name=team_name,
                rank_overall=rank,
                total_points=int(total_points),
                round_scores=round_scores,
            )
            team_scores.append(team_score)

        return rounds, team_scores
