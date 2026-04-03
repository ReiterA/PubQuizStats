"""Create sample Excel file for testing."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path

def create_sample_quiz_excel():
    """Create a sample Excel file with quiz data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz Results"

    # Add title
    ws['A1'] = "Spring Trivia Night 2024"
    ws['A1'].font = Font(size=14, bold=True)

    # Add metadata
    ws['A2'] = "Date: 2024-03-15"
    ws['A3'] = "Location: The Pub"

    # Add headers
    ws['A5'] = "Team Name"
    ws['B5'] = "Round 1"
    ws['C5'] = "Round 2"
    ws['D5'] = "Round 3"
    ws['E5'] = "Total"
    ws['F5'] = "Rank"

    # Format headers
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}5'].fill = header_fill
        ws[f'{col}5'].font = header_font
        ws[f'{col}5'].alignment = Alignment(horizontal='center')

    # Add sample data
    data = [
        ["London Eagles", 18, 16, 20, 54, 1],
        ["Quiz Masters", 17, 18, 18, 53, 2],
        ["Brain Busters", 15, 14, 16, 45, 3],
        ["Team Alpha", 14, 15, 14, 43, 4],
        ["The Thinkers", 12, 13, 11, 36, 5],
        ["Quiz Legends", 16, 17, 15, 48, 6],
    ]

    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if col_idx > 1:  # Numbers alignment
                cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 18
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 12

    # Save file
    output_path = Path(__file__).parent / "sample_quiz.xlsx"
    wb.save(output_path)
    print(f"Created sample Excel file: {output_path}")


if __name__ == "__main__":
    create_sample_quiz_excel()
